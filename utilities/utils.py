from openpyxl import Workbook, load_workbook
import csv
import softest
import logging
import inspect

class Utils(softest.TestCase):
    def assert_list_item_text(self, list_of_results, value):
        for stop_item in list_of_results:
            text = stop_item.text.strip()
            print(f"The text is: {text}")

            if text == value:
                print("test passed")
            else:
                print("test failed")

            self.soft_assert(self.assertEqual, value, text)

        self.assert_all()

    def custom_logger(logLevel = logging.DEBUG):
        # Set class/method name from where it's called.
        logger_name = inspect.stack()[1][3]

        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)

        # create file handler and set the log level with overwrite mode ('w')
        fh = logging.FileHandler(filename="automation.log") # fh = file handler

        # create formatter - how you want your logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s - %(message)s', datefmt = '%d/%m/%Y %I:%M:%S %p')

        # add formatter to console or file handler
        fh.setFormatter(formatter)

        # add console handler to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct+1):
            row = []
            for j in range(1, col_ct+1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        datalist = []
        csvdata = open(filename, "r")
        reader = csv.reader(csvdata)
        #Skip header
        next(reader)
        #Add CSV rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist




